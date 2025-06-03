import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext

def compare_and_highlight_excel(file1_path, file2_path, output_dir="highlighted_excel_files", status_callback=None):
    """
    Compares two Excel files, finds matching cell values, and highlights them in yellow
    in newly generated output files. Additionally, it highlights shared numeric values in red.

    Args:
        file1_path (str): Path to the first Excel file.
        file2_path (str): Path to the second Excel file.
        output_dir (str): Directory where the highlighted Excel files will be saved.
                          Defaults to "highlighted_excel_files".
        status_callback (callable, optional): A function to call with status updates.
                                              Defaults to None.
    """
    def update_status(message):
        if status_callback:
            status_callback(message)
        else:
            print(message)

    update_status("Starting comparison...")

    if not os.path.exists(file1_path):
        update_status(f"Error: File not found at {file1_path}")
        messagebox.showerror("Error", f"File not found at {file1_path}")
        return
    if not os.path.exists(file2_path):
        update_status(f"Error: File not found at {file2_path}")
        messagebox.showerror("Error", f"File not found at {file2_path}")
        return

    os.makedirs(output_dir, exist_ok=True)

    # Define the yellow fill style for all matches
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Define the red fill style for shared numbers
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    try:
        # Load all sheets from both Excel files into dictionaries of DataFrames
        update_status(f"Loading {file1_path}...")
        dfs1 = pd.read_excel(file1_path, sheet_name=None)
        update_status(f"Loading {file2_path}...")
        dfs2 = pd.read_excel(file2_path, sheet_name=None)
    except Exception as e:
        update_status(f"Error loading Excel files: {e}")
        messagebox.showerror("Error", f"Error loading Excel files: {e}")
        return

    # Prepare output file paths
    base_name1 = os.path.basename(file1_path).replace(".xlsx", "_highlighted.xlsx")
    base_name2 = os.path.basename(file2_path).replace(".xlsx", "_highlighted.xlsx")
    output_file1_path = os.path.join(output_dir, base_name1)
    output_file2_path = os.path.join(output_dir, base_name2)

    # Save initial DataFrames to new Excel files to prepare for openpyxl
    update_status(f"Saving initial DataFrames to {output_file1_path} and {output_file2_path}...")
    try:
        with pd.ExcelWriter(output_file1_path, engine='openpyxl') as writer:
            for sheet_name, df in dfs1.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        with pd.ExcelWriter(output_file2_path, engine='openpyxl') as writer:
            for sheet_name, df in dfs2.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        update_status(f"Error saving initial DataFrames: {e}")
        messagebox.showerror("Error", f"Error saving initial DataFrames: {e}")
        return

    # Load workbooks for highlighting
    update_status("Loading workbooks for highlighting...")
    try:
        wb1 = load_workbook(output_file1_path)
        wb2 = load_workbook(output_file2_path)
    except Exception as e:
        update_status(f"Error loading workbooks for highlighting: {e}")
        messagebox.showerror("Error", f"Error loading workbooks for highlighting: {e}")
        return

    update_status("Comparing and highlighting matches...")
    # Iterate through each sheet in the first workbook
    for sheet_name1, df1 in dfs1.items():
        if sheet_name1 not in wb1.sheetnames:
            update_status(f"Warning: Sheet '{sheet_name1}' not found in {output_file1_path}. Skipping.")
            continue
        ws1 = wb1[sheet_name1]

        # Iterate through each sheet in the second workbook
        for sheet_name2, df2 in dfs2.items():
            if sheet_name2 not in wb2.sheetnames:
                update_status(f"Warning: Sheet '{sheet_name2}' not found in {output_file2_path}. Skipping.")
                continue
            ws2 = wb2[sheet_name2]

            # Convert DataFrames to sets of values for efficient lookup
            # Flatten the DataFrame values into a single set for quick checking across sheets
            values1_set = set(df1.stack().dropna().astype(str).tolist())
            values2_set = set(df2.stack().dropna().astype(str).tolist())

            # Find common values between the two sheets (for yellow highlighting)
            common_values = values1_set.intersection(values2_set)

            # Find common numeric values between the two sheets (for red highlighting)
            common_numeric_values = set()
            for val in common_values:
                try:
                    # Attempt to convert to float to check if it's numeric
                    float(val)
                    common_numeric_values.add(val)
                except ValueError:
                    # Not a number, skip
                    pass

            if not common_values:
                continue # No common values between these two sheets, move to next pair

            # Highlight matches in the first workbook (ws1)
            for r_idx, row in enumerate(df1.values):
                for c_idx, cell_value in enumerate(row):
                    if pd.notna(cell_value):
                        cell_str_value = str(cell_value)
                        target_cell = ws1.cell(row=r_idx + 2, column=c_idx + 1)
                        if cell_str_value in common_numeric_values:
                            target_cell.fill = red_fill
                        elif cell_str_value in common_values:
                            target_cell.fill = yellow_fill

            # Highlight matches in the second workbook (ws2)
            for r_idx, row in enumerate(df2.values):
                for c_idx, cell_value in enumerate(row):
                    if pd.notna(cell_value):
                        cell_str_value = str(cell_value)
                        target_cell = ws2.cell(row=r_idx + 2, column=c_idx + 1)
                        if cell_str_value in common_numeric_values:
                            target_cell.fill = red_fill
                        elif cell_str_value in common_values:
                            target_cell.fill = yellow_fill

    # Save the highlighted workbooks
    try:
        wb1.save(output_file1_path)
        wb2.save(output_file2_path)
        final_message = (f"Comparison complete. Highlighted files saved to:\n"
                         f"- {output_file1_path}\n"
                         f"- {output_file2_path}")
        update_status(final_message)
        messagebox.showinfo("Success", final_message)
    except Exception as e:
        update_status(f"Error saving highlighted files: {e}")
        messagebox.showerror("Error", f"Error saving highlighted files: {e}")


class ExcelComparatorApp:
    def __init__(self, master):
        self.master = master
        master.title("ROYAL CLASSIFICATION SOCIETY") # Changed title here
        master.geometry("600x450") # Set a default window size
        master.resizable(False, False) # Make window not resizable

        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()

        # Frame for file selection
        file_frame = tk.Frame(master, padx=10, pady=10)
        file_frame.pack(pady=10)

        tk.Label(file_frame, text="Excel File 1:").grid(row=0, column=0, sticky="w", pady=5)
        self.entry1 = tk.Entry(file_frame, textvariable=self.file1_path, width=50)
        self.entry1.grid(row=0, column=1, padx=5, pady=5)
        tk.Button(file_frame, text="Browse", command=lambda: self.browse_file(self.file1_path)).grid(row=0, column=2, padx=5, pady=5)

        tk.Label(file_frame, text="Excel File 2:").grid(row=1, column=0, sticky="w", pady=5)
        self.entry2 = tk.Entry(file_frame, textvariable=self.file2_path, width=50)
        self.entry2.grid(row=1, column=1, padx=5, pady=5)
        tk.Button(file_frame, text="Browse", command=lambda: self.browse_file(self.file2_path)).grid(row=1, column=2, padx=5, pady=5)

        # Compare button
        self.compare_button = tk.Button(master, text="Compare and Highlight", command=self.run_comparison,
                                        bg="#4CAF50", fg="white", font=("Arial", 12, "bold"),
                                        relief="raised", bd=3, padx=10, pady=5)
        self.compare_button.pack(pady=20)

        # Status text area
        tk.Label(master, text="Status:").pack(anchor="w", padx=10)
        self.status_text = scrolledtext.ScrolledText(master, wrap=tk.WORD, width=70, height=10, font=("Arial", 9))
        self.status_text.pack(padx=10, pady=5)
        self.status_text.insert(tk.END, "Ready to compare Excel files.\n")
        self.status_text.config(state=tk.DISABLED) # Make it read-only

    def browse_file(self, path_var):
        """Opens a file dialog and updates the given StringVar with the selected file path."""
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if filepath:
            path_var.set(filepath)
            self.update_status_text(f"Selected: {filepath}\n")

    def update_status_text(self, message):
        """Appends a message to the status text area."""
        self.status_text.config(state=tk.NORMAL) # Enable for editing
        self.status_text.insert(tk.END, message)
        self.status_text.see(tk.END) # Scroll to the end
        self.status_text.config(state=tk.DISABLED) # Disable again

    def run_comparison(self):
        """Retrieves file paths and starts the comparison process."""
        file1 = self.file1_path.get()
        file2 = self.file2_path.get()

        if not file1 or not file2:
            messagebox.showwarning("Missing Files", "Please select both Excel files.")
            self.update_status_text("Error: Please select both Excel files.\n")
            return

        self.update_status_text("Comparison initiated...\n")
        # Run the comparison in a separate thread if it were a long-running task
        # For simplicity here, it runs directly, which might freeze the GUI for large files.
        # For very large files, consider using threading.
        compare_and_highlight_excel(file1, file2, status_callback=self.update_status_text)


if __name__ == "__main__":
    # Create dummy Excel files for testing if they don't exist
    excel_file1_dummy = "file1.xlsx"
    excel_file2_dummy = "file2.xlsx"

    if not os.path.exists(excel_file1_dummy) or not os.path.exists(excel_file2_dummy):
        print("Creating dummy Excel files for demonstration...")
        data1 = {
            'ID': [1, 2, 3, 4, 5],
            'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Eve'],
            'City': ['New York', 'London', 'Paris', 'New York', 'Tokyo'],
            'Value': [100, 200, 300, 400, 500] # Added for numeric comparison
        }
        df_test1 = pd.DataFrame(data1)
        df_test1.to_excel(excel_file1_dummy, index=False)

        data2 = {
            'Product': ['A', 'B', 'C', 'D', 'E'],
            'Value': [100, 250, 300, 450, 500], # Changed some values for better testing
            'City': ['London', 'Berlin', 'Paris', 'Rome', 'New York'],
            'Quantity': [10, 20, 30, 40, 50] # Added for numeric comparison
        }
        df_test2 = pd.DataFrame(data2)
        df_test2.to_excel(excel_file2_dummy, index=False)
        print("Dummy files created.")

    root = tk.Tk()
    app = ExcelComparatorApp(root)
    root.mainloop()